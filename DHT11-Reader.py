#!/usr/bin/python3

import Adafruit_DHT
import time
import datetime
import requests
import os
from openpyxl import Workbook

file_path = os.getenv('file','dump.xls')
url = os.getenv('url','127.0.0.1:8000/entries')
id = os.getenv('id','default')
pin = os.getenv('pin',23)
sleep_time = os.getenv('sleep',300)

print(file_path)
print(url)
print(id)
print(pin)
print(sleep_time)

current_row = 2
sensor = Adafruit_DHT.DHT11
wb = Workbook()
ws = wb.active
ws.cell(row=1,column=1,value='Date/Time')
ws.cell(row=1,column=2,value='Humidity')
ws.cell(row=1,column=3,value='Temperature')



while True:
    try:
        humidity, temperature = Adafruit_DHT.read_retry(sensor, pin)
        if humidity is not None and temperature is not None:
            r = requests.post(url,data={'id': id,'time': datetime.datetime.now(), 'temp': temperature, 'hum': humidity})
            ws.cell(row=current_row, column=1,value=datetime.datetime.now())
            ws.cell(row=current_row, column=2,value=humidity)
            ws.cell(row=current_row, column=3,value=temperature)
            current_row += 1
            wb.save(file_path)
            time.sleep(sleep_time)
    except KeyboardInterrupt:
        wb.save(file_path)
        break
    except Exception as e:
        print(str(e))
        wb.save(file_path)
